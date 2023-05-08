import requests
from bs4 import BeautifulSoup
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import re
from collections import Counter
import pandas as pd

def getInfo(url, row):
    IMAGE_HEIGHT = 280
    IMAGE_WIDTH = 340

    response = requests.get(url) #sending request using the given url
    soup = BeautifulSoup(response.content, 'html.parser') #parsing request

    internalLinks = findInternalLinks(soup)
    print("Amount of links:", len(internalLinks))
    
    words = getWordList(soup)
    print("Amount of words:", len(words))

    symbols = ''.join(words) #getting a string out of wordlist
    rus = re.findall('[а-яё]', symbols) #finding RUS letter frequency in symbols str
    eng = re.findall('[a-z]', symbols) #finding ENG letter frequency in symbols str
    spec = re.findall("[@_!#$%^&*()<>?/|}{~:]", symbols) #finding special symbol frequency in symbols str

    print(symbols)
    print(words)
    fig = plt.figure()
    fig.canvas.manager.window.title('Histograms for ' + url)
    axs = fig.subplots(nrows=2, ncols=2)
    
    # Adding the histograms to the subplots
    symbol_histogram(rus).plot(kind='barh', ax=axs[0, 0], rot=0, color='black')
    axs[0, 0].set_xlabel('Frequency')
    axs[0, 0].set_ylabel('Symbols')
    axs[0, 0].set_title('Rus symbols')
    
    symbol_histogram(eng).plot(kind='barh', ax=axs[0, 1], rot=0, color='black')
    axs[0, 1].set_xlabel('Frequency')
    axs[0, 1].set_ylabel('Symbols')
    axs[0, 1].set_title('Eng symbols')
    
    symbol_histogram(spec).plot(kind='barh', ax=axs[1, 0], rot=0, color='black')
    axs[1, 0].set_xlabel('Frequency')
    axs[1, 0].set_ylabel('Symbols')
    axs[1, 0].set_title('Spec symbols')
    
    getWordLength(words=words).plot(kind='barh', ax=axs[1, 1], rot=0, color='black')
    axs[1, 1].set_xlabel('Frequency')
    axs[1, 1].set_ylabel("Word's length")
    axs[1, 1].set_title('Rus symbols')
    
    plt.subplots_adjust(wspace=0.3, hspace=0.3)
    plt.show()
    
    symbol_histogram(rus).plot(kind='barh', rot=0, color='black')
    plt.savefig("rus" + str(row) + ".png")
    imgRus = Image("rus" + str(row) + ".png") #e.g. rus1.png
    imgRus.height = IMAGE_HEIGHT
    imgRus.width = IMAGE_WIDTH
    plt.close()
    
    symbol_histogram(eng).plot(kind='barh', rot=0, color='black')
    plt.savefig("eng" + str(row) + ".png")
    imgEng = Image("eng" + str(row) + ".png") #e.g. eng1.png
    imgEng.height = IMAGE_HEIGHT
    imgEng.width = IMAGE_WIDTH
    plt.close()
    
    symbol_histogram(spec).plot(kind='barh', rot=0, color='black')
    plt.savefig("spec" + str(row) + ".png")
    imgSpec = Image("spec" + str(row) + ".png") #e.g. spec1.png
    imgSpec.height = IMAGE_HEIGHT
    imgSpec.width = IMAGE_WIDTH
    plt.close()
    
    getWordLength(words=words).plot(kind='barh', rot=0, color='black')
    plt.savefig("Word Length" + str(row) + ".png")
    img4 = Image("Word Length"+str(row) + ".png")
    img4.height = IMAGE_HEIGHT
    img4.width = IMAGE_WIDTH
    plt.close()
    
    addNewRowTo(ws, url, len(internalLinks), len(words), imgRus, imgEng, imgSpec, img4, row)


def findInternalLinks(soup: BeautifulSoup):
    return [
        a.get('href') for a in soup.find_all('a')
        if a.get('href') and a.get('href').startswith('/')]

def getWordList(soup:BeautifulSoup):
    s = soup.text
    s = s.replace('\n', ' ')
    s = s.split(' ')
    return list(filter(None, s))

def getWordLength(words: list):
    word_len = list(map(lambda x: len(x), words))
    print(word_len)
    dict_word_len = Counter(list(filter(lambda a: a <= 15, word_len)))
    df = pd.DataFrame.from_dict(dict_word_len, orient='index')
    return df

def addNewRowTo(ws, url, refs, words, img1, img2, img3, img4, row):
    ws['A'+str(row)] = url
    ws['B'+str(row)] = refs
    ws['C'+str(row)] = words
    ws.add_image(img1, str('D') + str(row))
    ws.add_image(img2, str('E') + str(row))
    ws.add_image(img3, str('F') + str(row))
    ws.add_image(img4, str('G') + str(row))
    ws.row_dimensions[row].height = int(240 * 0.8)


def setUpTheHeadOf(ws):
    ws["A1"] = "URL"
    ws.column_dimensions["A"].width = int(200 * 0.2) #setting worksheet column dimensions
    ws["B1"] = "Amount of links"
    ws.column_dimensions["B"].width = int(100 * 0.2)
    ws["C1"] = "Amount of words"
    ws.column_dimensions["C"].width = int(100 * 0.2)
    ws["D1"] = "Symbol frequency gistogramm RUS"
    ws.column_dimensions["D"].width = int(240 * 0.2)
    ws["E1"] = "Symbol frequency gistogramm ENG"
    ws.column_dimensions["E"].width = int(240 * 0.2)
    ws["F1"] = "Symbol frequency gistogramm SPECS"
    ws.column_dimensions["F"].width = int(240 * 0.2)
    ws["G1"] = "Word length frequency gistogramm"
    ws.column_dimensions["G"].width = int(240 * 0.2)

def symbol_histogram(symbol_type):
    symbol_counter = Counter(symbol_type) #creates an iterable counter dictionary
    df = pd.DataFrame.from_dict(symbol_counter, orient='index').sort_index() #extracting dataframe from the counter dictionary
    return df

wb = Workbook()
ws = wb.active #worksheet
setUpTheHeadOf(ws)
getInfo("https://www.gog.com/ru", 2)
getInfo("http://devilmayquake.com", 3)
getInfo("https://developer.alexanderklimov.ru/android/kotlin/", 4)
wb.save(filename="Lab3.xlsx")

